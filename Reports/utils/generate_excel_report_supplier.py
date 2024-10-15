

import openpyxl
from openpyxl.styles import Font
from io import BytesIO

def generate_excel_report_supplier(report_data):
    # إنشاء كائن Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receiving Report"

    # عنوان التقرير
    ws.merge_cells('A1:F1')
    ws['A1'] = "تقرير عمليات الوارد"
    ws['A1'].font = Font(size=14, bold=True)

    # إضافة عناوين الأعمدة
    ws['A3'] = "تاريخ الوارد"
    ws['B3'] = "رقم الوثيقة"
    ws['C3'] = "اسم المستفيد"
    ws['D3'] = "الكمية المصروفة"
    ws['E3'] = "اسم الموزع"
    ws['F3'] = "اسم المستلم"

    ws['A3'].font = Font(bold=True)
    ws['B3'].font = Font(bold=True)
    ws['C3'].font = Font(bold=True)
    ws['D3'].font = Font(bold=True)
    ws['E3'].font = Font(bold=True)
    ws['F3'].font = Font(bold=True)

    # الحصول على بيانات الشحن
    receivings = report_data['Receiving']
    
    # إضافة البيانات إلى Excel
    row = 4
    for receiving in receivings:
        ws.cell(row=row, column=1, value=receiving.receiving_date)
        ws.cell(row=row, column=2, value=receiving.document_number)
        ws.cell(row=row, column=3, value=receiving.supplier.full_name)
        ws.cell(row=row, column=4, value=receiving.imported_quantity)
        ws.cell(row=row, column=5, value=receiving.delivered_by_name)
        ws.cell(row=row, column=6, value=receiving.received_by_name)
        row += 1

    # تعديل عرض الأعمدة
    for column in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[column].width = 25

    # حفظ الملف في الذاكرة
    buffer = BytesIO()
    wb.save(buffer)
    wb.close()  # إغلاق الملف لتحرير الذاكرة
    buffer.seek(0)
    return buffer
