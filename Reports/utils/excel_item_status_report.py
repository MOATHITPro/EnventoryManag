import openpyxl
from openpyxl.styles import Font
from io import BytesIO

def generate_excel_item_status(report_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير حالة الصنف"

    # عنوان التقرير
    ws.merge_cells('A1:C1')
    ws['A1'] = "تقرير حالة الصنف في المخازن"
    ws['A1'].font = Font(size=14, bold=True)

    # معلومات الصنف
    ws['A3'] = "اسم الصنف:"
    ws['B3'] = report_data['Item'].name

    # عناوين الجدول
    ws['A5'] = "المخزن"
    ws['B5'] = "الوحدة"
    ws['C5'] = "الكميات المتوفرة"
    ws['A5'].font = Font(bold=True)
    ws['B5'].font = Font(bold=True)
    ws['C5'].font = Font(bold=True)

    # بيانات الجدول
    row = 6
    for stock in report_data['StockItems']:
        ws.cell(row=row, column=1, value=stock.warehouse.name)
        ws.cell(row=row, column=2, value=stock.unit)
        ws.cell(row=row, column=3, value=stock.current_quantity)
        row += 1

    # تعديل عرض الأعمدة
    for column in ['A', 'B', 'C']:
        ws.column_dimensions[column].width = 25

    # حفظ الملف في الذاكرة
    buffer = BytesIO()
    wb.save(buffer)
    wb.close()  # إغلاق الملف لتحرير الذاكرة
    buffer.seek(0)
    return buffer
