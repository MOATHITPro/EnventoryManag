import openpyxl
from openpyxl.styles import Font
from io import BytesIO
import openpyxl
from openpyxl.styles import Font
from io import BytesIO

def generate_excel_warehouse_status(report_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير حالة المخزن"

    # عنوان التقرير
    ws.merge_cells('A1:C1')
    ws['A1'] = "تقرير حالة المخزن"
    ws['A1'].font = Font(size=14, bold=True)

    # معلومات المخزن
    ws['A3'] = "اسم المخزن:"
    ws['B3'] = report_data['Warehouse'].name

    # عناوين الجدول
    ws['A5'] = "اسم الصنف"
    ws['B5'] = "الكميات المتوفرة"
    ws['A5'].font = Font(bold=True)
    ws['B5'].font = Font(bold=True)

    # بيانات الجدول
    row=5
    for stock_item in report_data['StockItems']:
        ws.cell(row=row, column=1, value=stock_item.item.name)  # اسم الصنف
        ws.cell(row=row, column=2, value=stock_item.current_quantity)  # الكميات المتوفرة
        row += 1

    # تعديل عرض الأعمدة
    for column in ['A', 'B']:
        ws.column_dimensions[column].width = 25

    # حفظ الملف في الذاكرة
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
